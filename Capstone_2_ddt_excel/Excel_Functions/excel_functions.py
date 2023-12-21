from openpyxl import load_workbook

class Excel_Functions:
   def __init__(self, excel_file_name, excel_sheet_name):
      self.file = excel_file_name
      self.sheet = excel_sheet_name
       

   # fetch the row count
   def row_count(self):
      workbook = load_workbook(self.file)
      sheet = workbook[self.sheet]
      return sheet.max_row


   # fetch the column count
   def column_count(self):
      workbook = load_workbook(self.file)
      sheet = workbook[self.sheet]
      return sheet.max_column


   # read data from excel file
   def read_data(self, row_number, column_number):
      workbook = load_workbook(self.file)
      sheet = workbook[self.sheet]
      return sheet.cell(row=row_number, column=column_number).value


   # write data into excel file
   def write_data(self, row_number, column_number, data):
      workbook = load_workbook(self.file)
      sheet = workbook[self.sheet]
      sheet.cell(row=row_number, column=column_number).value = data
      workbook.save(self.file)


   # Read excel data from Excel file Capstone_2 sheet_2
   def read_excel_data_p2_t1(self):
      p2_t1_datas_lst = []
      for row_num in range (2,3):
         # ["result_row","lg_username","expected_result"]
         p2_t1_data = []
         for i in range(0,3):
            t_data = self.read_data(2,i+1)
            p2_t1_data.append(t_data)
      p2_t1_datas_lst.append(p2_t1_data)

      print("P2_t1_datas : ",p2_t1_datas_lst )
      return p2_t1_datas_lst
   

   # Read excel data from Excel file Captone_2 Sheet_3
   def read_excel_data_p2_t2(self):
      p2_t2_datas_lst = []
      for row_num in range(2,3):
         p2_t2_data = []
         for i in range(0,12):
            t_data = self.read_data(2,i+1)
            p2_t2_data.append(t_data)
         p2_t2_datas_lst.append(p2_t2_data)

         print("P2_t2_datas : ",p2_t2_datas_lst)
         return p2_t2_datas_lst
      

   def read_excel_data_p2_t3(self):
      p2_t3_datas_lst = []
      for row_num in range(2,3):
         p2_t3_data = []
         for i in range(0,4):
            t_data = self.read_data(2,i+1)
            p2_t3_data.append(t_data)
         p2_t3_datas_lst.append(p2_t3_data)

      print("p2_t3_datas : ", p2_t3_datas_lst)
      return p2_t3_datas_lst


ef =Excel_Functions("Capstone_2.xlsx","Sheet2")
ef.read_excel_data_p2_t1()
ef =Excel_Functions("Capstone_2.xlsx","Sheet3")
ef.read_excel_data_p2_t2()
ef =Excel_Functions("Capstone_2.xlsx","Sheet4")
ef.read_excel_data_p2_t3()
